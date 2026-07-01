"""
统计两个 Excel 文件中"点位国标"的出现次数 + 对应的时间字段。

- 2028项目工单.xls：筛选 派单时间 >= 2026年1月 且 故障原因 == "设备离线"
  额外输出：该 国标 对应的所有"派单时间"（多个用换行分隔）
- 2082项目全部工单_20260701导出.xlsx：无需筛选，全部统计
  额外输出：该 国标 对应的所有"完成期限"（多个用换行分隔）

输出列：点位国标 | 文件名1出现次数 | 派单时间 | 文件名2出现次数 | 完成期限
"""

import pandas as pd
from pathlib import Path

# ── 路径 ──────────────────────────────────────────
DATA_DIR = Path(__file__).parent / "data"
FILE1 = DATA_DIR / "2028项目工单.xls"
FILE2 = DATA_DIR / "2082项目全部工单_20260701导出.xlsx"
OUTPUT = DATA_DIR / "国标编码统计结果_v2.xlsx"

# ── 读取 ──────────────────────────────────────────
df1 = pd.read_excel(FILE1)
df2 = pd.read_excel(FILE2)

# ── 列名（按索引取，避免编码问题）─────────────────
# 文件1：col[2]=派单时间, col[6]=故障原因, col[16]=点位国标
col_time1 = df1.columns[2]    # 派单时间
col_fault = df1.columns[6]    # 故障原因
col_gb1 = df1.columns[16]     # 点位国标

# 文件2：col[3]=国标编码, col[5]=完成期限
col_gb2 = df2.columns[3]      # 国标编码
col_time2 = df2.columns[5]    # 完成期限

# ── 文件1：筛选 + 按国标聚合 ──────────────────────
df1[col_time1] = pd.to_datetime(df1[col_time1], errors="coerce")
mask = (
    (df1[col_time1] >= pd.Timestamp("2026-01-01"))
    & (df1[col_fault] == "设备离线")
)
filtered1 = df1.loc[mask, [col_gb1, col_time1]].copy()
filtered1[col_gb1] = filtered1[col_gb1].astype(str).str.strip()
filtered1.dropna(subset=[col_gb1], inplace=True)

# 按国标聚合：次数 + 所有派单时间（按时间排序，换行分隔）
def join_times(series):
    """将一系列时间排序后用换行符拼接"""
    cleaned = pd.Series(series.dropna().sort_values())
    return "\n".join(t.strftime("%Y-%m-%d %H:%M:%S") for t in cleaned)

agg1 = filtered1.groupby(col_gb1).agg(
    count=(col_time1, "count"),
    times=(col_time1, join_times),
)

# ── 文件2：按国标聚合 ──────────────────────────────
filtered2 = df2[[col_gb2, col_time2]].copy()
filtered2[col_gb2] = filtered2[col_gb2].astype(str).str.strip()
filtered2.dropna(subset=[col_gb2], inplace=True)

# 完成期限转为 datetime，方便排序
filtered2[col_time2] = pd.to_datetime(filtered2[col_time2], errors="coerce")

agg2 = filtered2.groupby(col_gb2).agg(
    count=(col_time2, "count"),
    times=(col_time2, join_times),
)

# ── 合并为一张表 ──────────────────────────────────
all_gb = sorted(set(agg1.index) | set(agg2.index))

result = pd.DataFrame({
    "点位国标": all_gb,
    FILE1.name: [agg1.loc[gb, "count"] if gb in agg1.index else 0 for gb in all_gb],
    "派单时间": [agg1.loc[gb, "times"] if gb in agg1.index else "" for gb in all_gb],
    FILE2.name: [agg2.loc[gb, "count"] if gb in agg2.index else 0 for gb in all_gb],
    "完成期限": [agg2.loc[gb, "times"] if gb in agg2.index else "" for gb in all_gb],
})

# ── 写入 Excel，并设置时间列自动换行 ──────────────
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    result.to_excel(writer, index=False, sheet_name="统计结果")
    ws = writer.sheets["统计结果"]

    # 时间列索引：C 列(派单时间=3), E 列(完成期限=5)
    from openpyxl.styles import Alignment
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        row[0].alignment = wrap_align
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        row[0].alignment = wrap_align

    # 自动调整列宽（估算）
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

# ── 打印摘要 ──────────────────────────────────────
print(f"文件1（{FILE1.name}）筛选后行数: {len(filtered1)}")
print(f"文件2（{FILE2.name}）总行数: {len(filtered2)}")
print(f"去重后的点位国标总数: {len(result)}")
print(f"结果已保存到: {OUTPUT}")
print()
# 打印多条目示例
multi = result[result[FILE1.name] > 1].head(5)
if len(multi) > 0:
    print("=== 文件1 中有多个派单时间的示例 ===")
    for _, row in multi.iterrows():
        print(f"  {row['点位国标']} ({row[FILE1.name]}次):")
        for line in row["派单时间"].split("\n"):
            print(f"    - {line}")
