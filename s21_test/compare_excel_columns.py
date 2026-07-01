"""
Excel 字段对比工具
===================
批量分析多个 Excel 文件的字段（列名）是否一致，并生成一份可阅读的 Excel 分析报告。

用法:
    python compare_excel_columns.py              # 分析当前目录
    python compare_excel_columns.py ./data       # 分析指定目录

输出:
    字段对比报告.xlsx（包含 5 个 Sheet：字段汇总、字段矩阵、缺失字段、两两比较、一致性总结）
"""

from pathlib import Path
import sys
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# 配置（可按需修改）
# ============================================================

# False: 仅分析每个 Excel 的第一个 Sheet
# True:  分析所有 Sheet
READ_ALL_SHEETS = False

# 默认分析目录（命令行参数可覆盖）
DEFAULT_FOLDER = Path(".")

# 输出文件名
OUTPUT_FILE = "字段对比报告.xlsx"

# 支持的 Excel 扩展名
SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}

# ============================================================
# 工具函数
# ============================================================


def _display_name(file_name: str, sheet_name: str) -> str:
    """生成显示名称：仅分析首 Sheet 时只返回文件名，否则返回 '文件名|Sheet名'。"""
    if READ_ALL_SHEETS:
        return f"{file_name}|{sheet_name}"
    return file_name


def _get_sheet_columns(excel_path: Path, sheet_name: str) -> List[str]:
    """读取指定 Sheet 的表头（第一行），返回列名列表。"""
    df = pd.read_excel(excel_path, sheet_name=sheet_name, nrows=0)
    return list(df.columns)


# ============================================================
# 核心函数
# ============================================================


def load_excel_columns(folder: Path) -> Dict[Tuple[str, str], List[str]]:
    """
    加载目录下所有 Excel 文件的字段信息。

    返回:
        {(文件名, Sheet名): [字段列表], ...}

    每个 Excel 文件：
      - READ_ALL_SHEETS=False: 只读取第一个 Sheet
      - READ_ALL_SHEETS=True:  读取所有 Sheet
    跳过输出报告自身。
    """
    data: Dict[Tuple[str, str], List[str]] = {}

    for f in sorted(folder.iterdir()):
        if f.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        if f.name == OUTPUT_FILE:
            continue

        try:
            xf = pd.ExcelFile(f)

            if READ_ALL_SHEETS:
                for sheet_name in xf.sheet_names:
                    cols = _get_sheet_columns(f, sheet_name)
                    data[(f.name, sheet_name)] = cols
            else:
                first_sheet = xf.sheet_names[0]
                cols = _get_sheet_columns(f, first_sheet)
                data[(f.name, first_sheet)] = cols

        except Exception as e:
            print(f"[警告] 读取失败: {f.name} — {e}")

    return data


def build_field_summary(data: Dict[Tuple[str, str], List[str]]) -> pd.DataFrame:
    """
    Sheet1: 字段汇总
    列: Excel文件 | Sheet | 字段
    每行 = 一个文件/Sheet 的一个字段。
    """
    rows = []
    for (file_name, sheet_name), cols in data.items():
        for c in cols:
            rows.append([file_name, sheet_name, c])
    return pd.DataFrame(rows, columns=["Excel文件", "Sheet", "字段"])


def build_field_matrix(
    data: Dict[Tuple[str, str], List[str]], all_columns: List[str]
) -> pd.DataFrame:
    """
    Sheet2: 字段矩阵
    行 = 字段名称, 列 = 每个文件/Sheet
    存在 → ✔, 不存在 → ✘
    """
    matrix = pd.DataFrame({"字段名称": all_columns})

    for (file_name, sheet_name), cols in data.items():
        col_name = _display_name(file_name, sheet_name)
        col_set = set(cols)
        matrix[col_name] = ["✔" if c in col_set else "✘" for c in all_columns]

    return matrix


def build_missing_fields(
    data: Dict[Tuple[str, str], List[str]], all_columns: List[str]
) -> pd.DataFrame:
    """
    Sheet3: 缺失字段
    列: Excel文件 | Sheet | 缺失字段
    若某文件/Sheet 无缺失，该行显示 "无"。
    """
    ref_set = set(all_columns)
    rows = []
    for (file_name, sheet_name), cols in data.items():
        missing = sorted(ref_set - set(cols))
        label = "、".join(missing) if missing else "无"
        rows.append([file_name, sheet_name, label])
    return pd.DataFrame(rows, columns=["Excel文件", "Sheet", "缺失字段"])


def build_pairwise_comparison(
    data: Dict[Tuple[str, str], List[str]],
) -> pd.DataFrame:
    """
    Sheet4: 两两比较
    列: 文件A | 文件B | A独有字段 | B独有字段
    遍历所有 (文件/Sheet) 两两组合。
    """
    items = list(data.items())
    rows = []
    for i in range(len(items)):
        for j in range(i + 1, len(items)):
            (fa, sa), ca = items[i]
            (fb, sb), cb = items[j]

            name_a = _display_name(fa, sa)
            name_b = _display_name(fb, sb)

            set_a, set_b = set(ca), set(cb)
            only_a = "、".join(sorted(set_a - set_b)) or "无"
            only_b = "、".join(sorted(set_b - set_a)) or "无"

            rows.append([name_a, name_b, only_a, only_b])

    if not rows:
        rows.append(["无", "无", "无", "无"])

    return pd.DataFrame(rows, columns=["文件A", "文件B", "A独有字段", "B独有字段"])


def build_consistency_summary(
    data: Dict[Tuple[str, str], List[str]], all_columns: List[str]
) -> pd.DataFrame:
    """
    Sheet5: 一致性总结
    列: 文件 | Sheet | 字段数量 | 是否一致
    字段集合与全部字段的并集一致 → ✔, 否则 → ✘
    """
    ref_set = set(all_columns)
    rows = []
    for (file_name, sheet_name), cols in data.items():
        ok = "✔" if set(cols) == ref_set else "✘"
        rows.append([file_name, sheet_name, len(cols), ok])
    return pd.DataFrame(rows, columns=["文件", "Sheet", "字段数量", "是否一致"])


# ============================================================
# 样式与导出
# ============================================================


def _apply_styles(worksheet):
    """为工作表应用统一的格式：冻结首行、自动筛选、自适应列宽、表头加粗。"""
    header_fill = PatternFill("solid", fgColor="DDDDDD")

    # 表头样式
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 冻结首行 + 自动筛选
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    # 自适应列宽（上限 42 字符）
    for col_cells in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        worksheet.column_dimensions[col_letter].width = min(max_len + 3, 42)


def export_report(
    data: Dict[Tuple[str, str], List[str]],
    output_path: Path,
) -> None:
    """将分析结果写入 Excel 报告（5 个 Sheet）。"""
    # 全部字段（排序，作为比较基准）
    all_columns = sorted({c for cols in data.values() for c in cols})

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet1: 字段汇总
        build_field_summary(data).to_excel(
            writer, sheet_name="字段汇总", index=False
        )

        # Sheet2: 字段矩阵（最重要）
        build_field_matrix(data, all_columns).to_excel(
            writer, sheet_name="字段矩阵", index=False
        )

        # Sheet3: 缺失字段
        build_missing_fields(data, all_columns).to_excel(
            writer, sheet_name="缺失字段", index=False
        )

        # Sheet4: 两两比较
        build_pairwise_comparison(data).to_excel(
            writer, sheet_name="两两比较", index=False
        )

        # Sheet5: 一致性总结
        build_consistency_summary(data, all_columns).to_excel(
            writer, sheet_name="一致性总结", index=False
        )

        # 统一套用样式
        for ws in writer.book.worksheets:
            _apply_styles(ws)

    print(f"[完成] 已生成: {output_path.resolve()}")


# ============================================================
# 入口
# ============================================================


def main(folder: Path | None = None) -> None:
    """主流程：加载 → 构建报告 → 导出。"""
    if folder is None:
        folder = DEFAULT_FOLDER

    if not folder.exists() or not folder.is_dir():
        print(f"[错误] 目录不存在: {folder}")
        sys.exit(1)

    print(f"[分析目录] {folder.resolve()}")
    print(f"[模式] {'所有 Sheet' if READ_ALL_SHEETS else '仅第一个 Sheet'}")

    data = load_excel_columns(folder)

    if not data:
        print("[错误] 未找到任何 Excel 文件")
        sys.exit(1)

    file_count = len({f for (f, _) in data.keys()})
    entry_count = len(data)
    print(f"[信息] 已加载 {file_count} 个文件" + (
        f"（{entry_count} 个 Sheet）" if READ_ALL_SHEETS else ""
    ))

    output_path = folder / OUTPUT_FILE
    export_report(data, output_path)


if __name__ == "__main__":
    # 支持命令行指定目录: python compare_excel_columns.py ./some-folder
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FOLDER
    main(target)
